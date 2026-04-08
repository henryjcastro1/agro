"""
Generador de archivos Excel a partir de plantillas
"""
import os
import shutil
from openpyxl import load_workbook
from config import MARCA_SI, PLANTILLA_CHECKLIST, PLANTILLA_HOJA_VIDA
from utils import generar_nombre_archivo

def escribir_en_celda(ws, coordenada, valor):
    """Escribe en una celda manejando celdas fusionadas"""
    try:
        ws[coordenada] = valor
        return True
    except Exception as e:
        print(f"    ❌ Error en {coordenada}: {e}")
        return False

def generar_checklist(datos, output_path):
    """Genera el archivo Excel del checklist"""
    print(f"\n{'='*50}")
    print(f"📝 Generando Checklist...")
    print(f"{'='*50}")
    
    shutil.copy(PLANTILLA_CHECKLIST, output_path)
    wb = load_workbook(output_path)
    ws = wb["Mantenimiento"]
    
    try:
        # ========== DATOS BÁSICOS ==========
        print("\n📋 DATOS BÁSICOS:")
        escribir_en_celda(ws, "C5", datos["cliente"])        # Cliente
        escribir_en_celda(ws, "C6", datos["fecha"])          # Fecha
        escribir_en_celda(ws, "C7", datos["ciudad"])         # Ciudad
        escribir_en_celda(ws, "C8", datos["area"])           # Área
        escribir_en_celda(ws, "C9", datos["nombre_equipo"])  # Nombre equipo
        
        # ========== DATOS USUARIO ==========
        print("\n👤 DATOS USUARIO:")
        escribir_en_celda(ws, "H6", datos["nombre_usuario"])  # Nombre usuario
        escribir_en_celda(ws, "H7", datos["cedula"])          # Cédula
        escribir_en_celda(ws, "H8", datos["cargo"])           # Cargo
        escribir_en_celda(ws, "H9", datos["nombre_soporte"])  # Nombre soporte
        
        # ========== ESPECIFICACIONES TÉCNICAS ==========
        print("\n🔧 ESPECIFICACIONES TÉCNICAS:")
        escribir_en_celda(ws, "A35", datos["num_activo"])     # Número de activo
        escribir_en_celda(ws, "C35", datos["serial"])         # Serial
        escribir_en_celda(ws, "D35", datos["marca"])          # Marca
        escribir_en_celda(ws, "H34", datos["modelo"])         # Modelo
        
        # ========== HARDWARE ==========
        print("\n💻 HARDWARE:")
        escribir_en_celda(ws, "A37", datos["procesador"])     # Procesador
        escribir_en_celda(ws, "E36", datos["ram"])            # RAM
        escribir_en_celda(ws, "E37", datos["disco_duro"])     # Disco duro
        
        # ========== VERIFICACIÓN INICIAL ==========
        print("\n📋 VERIFICACIÓN INICIAL:")
        verificaciones_ini = {
            "Encender estación": "C22",
            "Arranque del SO": "C23",
            "Disco Duro": "C24",
            "CD Rom y/o DVD": "C25",
            "Monitor": "C26",
            "Mouse": "C27",
            "Teclado": "C28"
        }
        for item, celda in verificaciones_ini.items():
            estado = datos["verif_inicial"].get(item, {})
            texto = estado.get("estado", "")
            if estado.get("falla"):
                texto += f" - {estado['falla']}"
            escribir_en_celda(ws, celda, texto)
            print(f"    ✅ {item} → {celda}: {texto[:40]}")
        
        # ========== MANTENIMIENTO FÍSICO ==========
        print("\n🔧 MANTENIMIENTO FÍSICO:")
        fisico_celdas = {
            "Remover polvo interior": "C40",
            "Verificar conexiones internas": "C41",
            "Cerrar CPU y limpiar carcaza": "C42",
            "Conectar Cables de Potencia y Periféricos": "C43"
        }
        for actividad, celda in fisico_celdas.items():
            if datos["mtto_fisico"].get(actividad, False):
                escribir_en_celda(ws, celda, MARCA_SI)
                print(f"    ✅ {actividad} → {celda}: ✓")
            else:
                escribir_en_celda(ws, celda, "")
                print(f"    ⭕ {actividad} → {celda}: vacío")
        
        # ========== LIMPIEZA DE PERIFÉRICOS ==========
        print("\n🧹 LIMPIEZA DE PERIFÉRICOS:")
        perifericos_celdas = {
            "Limpiar teclado": "F40",
            "Limpiar Monitor": "F41",
            "Limpiar Mouse": "F42"
        }
        for actividad, celda in perifericos_celdas.items():
            if datos["mtto_fisico"].get(actividad, False):
                escribir_en_celda(ws, celda, MARCA_SI)
                print(f"    ✅ {actividad} → {celda}: ✓")
            else:
                escribir_en_celda(ws, celda, "")
                print(f"    ⭕ {actividad} → {celda}: vacío")
        
        # ========== MANTENIMIENTO LÓGICO ==========
        print("\n💾 MANTENIMIENTO LÓGICO:")
        logico_celdas = {
            "Limpieza de archivos temporales": "C46",
            "Análisis rápido con el antivirus del PC": "C47",
            "Actualizaciones de Windows": "C48"
        }
        for actividad, celda in logico_celdas.items():
            if datos["mtto_logico"].get(actividad, False):
                escribir_en_celda(ws, celda, MARCA_SI)
                print(f"    ✅ {actividad} → {celda}: ✓")
            else:
                escribir_en_celda(ws, celda, "")
                print(f"    ⭕ {actividad} → {celda}: vacío")
        
        # ========== VERIFICACIÓN FINAL ==========
        print("\n✅ VERIFICACIÓN FINAL:")
        verificaciones_fin = {
            "Encender estación": "C52",
            "Arranque del SO": "C53",
            "Disco Duro": "C54",
            "CD Rom y/o DVD": "C55",
            "Monitor": "C56",
            "Mouse": "C57",
            "Teclado": "C58"
        }
        for item, celda in verificaciones_fin.items():
            estado = datos["verif_final"].get(item, {})
            texto = estado.get("estado", "")
            if estado.get("falla"):
                texto += f" - {estado['falla']}"
            escribir_en_celda(ws, celda, texto)
            print(f"    ✅ {item} → {celda}: {texto[:40]}")
        
        # ========== TIPO DE EQUIPO ==========
        print("\n🖥️ TIPO DE EQUIPO:")
        tipo = datos["tipo_equipo"]
        # Limpiar ambas celdas primero
        escribir_en_celda(ws, "B32", "")
        escribir_en_celda(ws, "D32", "")
        if tipo == "PC de escritorio":
            escribir_en_celda(ws, "B32", MARCA_SI)
            print(f"    ✅ PC de escritorio → B32: ✓")
        elif tipo == "Portatil":
            escribir_en_celda(ws, "D32", MARCA_SI)
            print(f"    ✅ Portátil → D32: ✓")
        
        # ========== OBSERVACIONES ==========
        print("\n📝 OBSERVACIONES:")
        escribir_en_celda(ws, "A59", datos.get("obs_checklist", "N/A"))
        
    except Exception as e:
        print(f"❌ Error generando checklist: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    wb.save(output_path)
    print(f"\n✅ Checklist guardado: {output_path}")

def generar_hoja_vida(datos, output_path):
    """Genera el archivo Excel de la hoja de vida"""
    print(f"\n{'='*50}")
    print(f"📝 Generando Hoja de Vida...")
    print(f"{'='*50}")
    
    shutil.copy(PLANTILLA_HOJA_VIDA, output_path)
    wb = load_workbook(output_path)
    ws = wb["Mantenimiento"]
    
    try:
        # ========== DATOS BÁSICOS ==========
        print("\n📋 DATOS BÁSICOS:")
        escribir_en_celda(ws, "B18", datos["nombre_equipo"])
        escribir_en_celda(ws, "E19", datos["cliente"])
        escribir_en_celda(ws, "B20", datos["ciudad"])
        escribir_en_celda(ws, "B21", datos["fecha"])
        
        # ========== ESPECIFICACIONES ==========
        print("\n🔧 ESPECIFICACIONES:")
        escribir_en_celda(ws, "A30", datos["num_activo"])
        escribir_en_celda(ws, "C30", datos["serial"])
        escribir_en_celda(ws, "F30", datos["marca"])
        escribir_en_celda(ws, "H30", datos["modelo"])
        
        # ========== HARDWARE ==========
        print("\n💻 HARDWARE:")
        escribir_en_celda(ws, "A32", datos["procesador"])
        escribir_en_celda(ws, "E32", datos["ram"])
        escribir_en_celda(ws, "G32", datos["disco_duro"])
        
        # ========== TIPO DE EQUIPO ==========
        print("\n🖥️ TIPO DE EQUIPO:")
        tipo = datos["tipo_equipo"]
        escribir_en_celda(ws, "B26", "")
        escribir_en_celda(ws, "D26", "")
        escribir_en_celda(ws, "F26", "")
        
        if tipo == "PC de escritorio":
            escribir_en_celda(ws, "B26", MARCA_SI)
            print(f"    ✅ PC de escritorio → B26: ✓")
        elif tipo == "Portatil":
            escribir_en_celda(ws, "D26", MARCA_SI)
            print(f"    ✅ Portátil → D26: ✓")
        elif tipo == "Impresora":
            escribir_en_celda(ws, "F26", MARCA_SI)
            print(f"    ✅ Impresora → F26: ✓")
        
        # ========== ESTADO DEL EQUIPO ==========
        print("\n📊 ESTADO DEL EQUIPO:")
        estado = datos["estado_equipo"]
        escribir_en_celda(ws, "B37", "")
        escribir_en_celda(ws, "D37", "")
        escribir_en_celda(ws, "F37", "")
        
        if estado == "Bueno":
            escribir_en_celda(ws, "B37", MARCA_SI)
            print(f"    ✅ Estado Bueno → B37: ✓")
        elif estado == "Regular":
            escribir_en_celda(ws, "D37", MARCA_SI)
            print(f"    ✅ Estado Regular → D37: ✓")
        elif estado == "Malo":
            escribir_en_celda(ws, "F37", MARCA_SI)
            print(f"    ✅ Estado Malo → F37: ✓")
        
        # ========== TIPO DE MANTENIMIENTO ==========
        print("\n🛠️ TIPO DE MANTENIMIENTO:")
        tipo_mtto = datos["tipo_mtto"]
        escribir_en_celda(ws, "B42", "")
        escribir_en_celda(ws, "D42", "")
        escribir_en_celda(ws, "F42", "")
        
        if tipo_mtto == "Preventivo":
            escribir_en_celda(ws, "B42", MARCA_SI)
            print(f"    ✅ Preventivo → B42: ✓")
        elif tipo_mtto == "Correctivo":
            escribir_en_celda(ws, "D42", MARCA_SI)
            print(f"    ✅ Correctivo → D42: ✓")
        elif tipo_mtto == "Instalación":
            escribir_en_celda(ws, "F42", MARCA_SI)
            print(f"    ✅ Instalación → F42: ✓")
        
        # ========== OBSERVACIONES ==========
        print("\n📝 OBSERVACIONES:")
        escribir_en_celda(ws, "A45", datos.get("obs_hoja_vida", "N/A"))
        
    except Exception as e:
        print(f"❌ Error generando hoja de vida: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    wb.save(output_path)
    print(f"\n✅ Hoja de Vida guardada: {output_path}")

def generar_archivos_excel(datos, carpeta_salida):
    """Genera ambos archivos Excel y devuelve sus rutas"""
    nombre_base = generar_nombre_archivo(datos, "")
    
    ruta_checklist = os.path.join(carpeta_salida, f"Checklist_{nombre_base}.xlsx")
    ruta_hoja_vida = os.path.join(carpeta_salida, f"HojaDeVida_{nombre_base}.xlsx")
    
    generar_checklist(datos, ruta_checklist)
    generar_hoja_vida(datos, ruta_hoja_vida)
    
    return {
        'checklist': ruta_checklist,
        'hoja_vida': ruta_hoja_vida
    }